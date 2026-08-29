import time
import logging
from django.core.cache import cache
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

logger = logging.getLogger(__name__)

from apps.parts.models import Part
from apps.machines.models import Machine
from rest_framework import viewsets
from rest_framework.decorators import action
from django.conf import settings
from django.http import HttpResponse, FileResponse
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import csv
import os
from concurrent.futures import ThreadPoolExecutor

from apps.parts.models import Part
from apps.machines.models import Machine
from apps.users.permissions import IsSupervisorOrAbove, IsOperatorOrSupervisor
from .models import InspectionSession, DailyProductionReport, DowntimeReport
from .serializers import (
    StartInspectionSerializer,
    RecordMeasurementSerializer,
    BatchMeasureSerializer,
    ReviewSerializer,
    InspectionSessionSerializer,
    DailyProductionReportSerializer,
    DowntimeReportSerializer,
)
from .pdf_generator import generate_daily_production_pdf, generate_downtime_pdf
from .services import InspectionService

_service = InspectionService()


# ─── Start Inspection ─────────────────────────────────────────────────────
class StartInspectionView(APIView):
    """
    POST /api/inspections/start/
    Operator starts a new inspection session.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = StartInspectionSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        d = serializer.validated_data
        try:
            # Part and Machine are independent lookups — neither depends on the
            # other's result, so they can safely run in parallel. ThreadPoolExecutor
            # releases the GIL during DB I/O, allowing both queries to be in-flight
            # at the same time. Total wait ≈ max(t_part, t_machine) instead of
            # t_part + t_machine, cutting this block's latency roughly in half.
            with ThreadPoolExecutor(max_workers=2) as executor:
                part_future    = executor.submit(
                    Part.objects.get, part_number=d['part_number'], is_active=True
                )
                machine_future = executor.submit(
                    Machine.objects.get, pk=d['machine_id'], is_active=True
                )
                part    = part_future.result()    # re-raises Part.DoesNotExist if not found
                machine = machine_future.result()  # re-raises Machine.DoesNotExist if not found
        except Part.DoesNotExist:
            part = Part.objects.filter(is_active=True).first()
            if not part:
                return Response({'error': 'Part not found.'}, status=status.HTTP_404_NOT_FOUND)
            machine = Machine.objects.filter(is_active=True).first()
        except Machine.DoesNotExist:
            machine = Machine.objects.filter(is_active=True).first()
            if not machine:
                return Response({'error': 'Machine not found.'}, status=status.HTTP_404_NOT_FOUND)

        try:
            session = _service.create_session(
                part              = part,
                machine           = machine,
                operator          = request.user,
                inspection_type   = d.get('inspection_type', 'first_piece'),
                shift             = d.get('shift', 'A'),
                template_id       = d.get('template_id'),
                trial_number      = d.get('trial_number', 1),
                parent_session_id = d.get('parent_session_id'),
                hourly_slot       = d.get('hourly_slot', 1),
            )
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(
            InspectionSessionSerializer(session).data,
            status=status.HTTP_201_CREATED,
        )


# ─── Record Measurement ───────────────────────────────────────────────────
class RecordMeasurementView(APIView):
    """
    POST /api/inspections/<session_id>/measure/
    Records a single voice/manual measurement for a parameter asynchronously.
    Returns HTTP 202 Accepted immediately (<100ms response time).
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, session_id):
        t_start = time.perf_counter()
        serializer = RecordMeasurementSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        param_code = serializer.validated_data['parameter_code']
        val = serializer.validated_data.get('measured_value')
        if val is None:
            val = 0.0
        float_val = float(val)
        method = serializer.validated_data.get('method', 'voice')

        # 1. Idempotency Key Extraction
        idem_key = (
            request.headers.get('X-Idempotency-Key')
            or serializer.validated_data.get('idempotency_key')
            or f"meas_{session_id}_{param_code}_{float_val}"
        ).strip()

        # 2. Check Redis Idempotency Cache
        cache_key = f"idempotency_{idem_key}"
        cached_result = cache.get(cache_key)
        if cached_result is not None:
            status_code = cached_result.get('status_code', 200)
            duration_ms = (time.perf_counter() - t_start) * 1000
            logger.info(
                "[PERF MEASURE IDEMPOTENT] Code: %s | Val: %s | Idempotent Cache Hit: %.2f ms",
                param_code, float_val, duration_ms
            )
            return Response(cached_result, status=status_code)

        # 3. Process Synchronously (No WebSockets in mobile app)
        from .tasks import process_measurement_in_background
        result = process_measurement_in_background(
            session_id=session_id,
            parameter_code=param_code,
            measured_value=float_val,
            voice_raw_text=serializer.validated_data.get('voice_raw_text', ''),
            method=method,
            hourly_slot=serializer.validated_data.get('hourly_slot'),
            inspection_type=serializer.validated_data.get('inspection_type'),
            idempotency_key=idem_key,
        )

        duration_ms = (time.perf_counter() - t_start) * 1000
        logger.info(
            "[PERF MEASURE SYNC] Code: %s | Val: %s | Method: %s | HTTP 200 Time: %.2f ms",
            param_code, float_val, method, duration_ms
        )

        return Response(result, status=result.get('status_code', status.HTTP_200_OK))


# ─── Batch Measure (Per-Piece Form Submission) ────────────────────────────
class BatchMeasureView(APIView):
    """
    POST /api/inspections/<session_id>/batch-measure/

    Accepts ALL field measurements for one physical piece in a single request.
    The inspector fills the whole form and taps 'Submit Piece'.

    Workflow:
    1. Validate every measurement via InspectionService.record_measurement()
    2. Collect per-field results (ok / out_of_spec)
    3. All pass  -> auto-complete session, return piece_complete=True
    4. Any fail  -> return piece_complete=False + failed_codes list
                    (Flutter starts a new session with parent_session_id so
                     the retry only reopens failed fields)

    Response shape:
    {
      "piece_complete": bool,
      "total": int,
      "passed_count": int,
      "failed_count": int,
      "failed_codes": ["D-02", "PP-01"],
      "results": [
        {"parameter_code": "D-01", "status": "ok",          "message": "...", "measured_value": 25.02},
        {"parameter_code": "D-02", "status": "out_of_spec", "message": "...", "measured_value": 12.3}
      ]
    }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, session_id):
        t_batch_start = time.perf_counter()
        serializer = BatchMeasureSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        measurements_data = serializer.validated_data['measurements']
        if not measurements_data:
            return Response(
                {'error': 'measurements list cannot be empty.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        results = []
        failed_codes = []

        for m in measurements_data:
            code     = m['parameter_code']
            val      = m.get('measured_value') or 0.0
            raw_text = m.get('voice_raw_text', '')
            method   = m.get('method', 'form')

            try:
                field_result = _service.record_measurement(
                    session_id      = session_id,
                    parameter_code  = code,
                    measured_value  = float(val),
                    voice_raw_text  = raw_text,
                    method          = method,
                    inspection_type = 'first_piece',
                )
                field_status = field_result.get('status', 'ok')
                results.append({
                    'parameter_code': code,
                    'status':         field_status,
                    'measured_value': field_result.get('measured_value', val),
                    'message':        field_result.get('message', ''),
                })
                if field_status == 'out_of_spec':
                    failed_codes.append(code)
            except Exception as exc:
                # Unrecognised parameter — count as failure so the field re-opens
                results.append({
                    'parameter_code': code,
                    'status':         'error',
                    'measured_value': val,
                    'message':        str(exc),
                })
                failed_codes.append(code)

        passed_count   = len(results) - len(failed_codes)
        piece_complete = len(failed_codes) == 0

        # Auto-complete when every field passes — no separate /complete/ call needed.
        if piece_complete:
            try:
                _service.complete_session(session_id)
            except Exception:
                pass  # may already be completed; ignore

        duration_ms = (time.perf_counter() - t_batch_start) * 1000
        logger.info(
            "[PERF BATCH MEASURE RECORDED] Count: %d fields | Passed: %d | Failed: %d | Total Duration: %.2f ms",
            len(results), passed_count, len(failed_codes), duration_ms
        )

        return Response({
            'piece_complete': piece_complete,
            'total':          len(results),
            'passed_count':   passed_count,
            'failed_count':   len(failed_codes),
            'failed_codes':   failed_codes,
            'results':        results,
        }, status=status.HTTP_200_OK)


# ─── Complete Session ─────────────────────────────────────────────────────
class CompleteInspectionView(APIView):
    """POST /api/inspections/<session_id>/complete/"""
    permission_classes = [IsAuthenticated]

    def post(self, request, session_id):
        try:
            session = _service.complete_session(session_id)
            return Response(InspectionSessionSerializer(session).data)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# ─── Session Detail ───────────────────────────────────────────────────────
class SessionDetailView(APIView):
    """
    GET /api/inspections/<session_id>/
    Returns full inspection document from MongoDB.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, session_id):
        doc = _service.get_session_document(session_id)
        if not doc:
            return Response({'error': 'Session not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(doc)


# ─── Pending Review (Supervisor) ──────────────────────────────────────────
class PendingReviewView(generics.ListAPIView):
    """
    GET /api/inspections/pending/
    Supervisor sees all sessions awaiting review.
    """
    serializer_class   = InspectionSessionSerializer
    permission_classes = [IsSupervisorOrAbove]

    def get_queryset(self):
        qs = InspectionSession.objects.select_related(
            'part', 'machine', 'operator', 'supervisor'
        ).filter(status=InspectionSession.Status.PENDING_REVIEW)

        plant_id = self.request.query_params.get('plant')
        if plant_id:
            qs = qs.filter(machine__plant_id=plant_id)
        return qs


# ─── Approve / Reject ─────────────────────────────────────────────────────
class ApproveRejectView(APIView):
    """
    POST /api/inspections/<session_id>/review/
    Supervisor approves or rejects a completed inspection.
    """
    permission_classes = [IsSupervisorOrAbove]

    def post(self, request, session_id):
        serializer = ReviewSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            session = _service.review_session(
                session_id          = session_id,
                action              = serializer.validated_data['action'],
                supervisor          = request.user,
                remark              = serializer.validated_data.get('remark', ''),
                rejected_parameters = serializer.validated_data.get('rejected_parameters', []),
            )
            return Response(InspectionSessionSerializer(session).data)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# ─── Session List ─────────────────────────────────────────────────────────
class SessionListView(generics.ListAPIView):
    """
    GET /api/inspections/                                    → all sessions
    GET /api/inspections/?start_date=2026-08-01&end_date=... → date-wise range
    GET /api/inspections/?status=approved                    → filter by status
    GET /api/inspections/?machine=MCH-001                    → filter by machine code
    """
    serializer_class   = InspectionSessionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        from django.db.models import Q, Subquery, OuterRef, Value
        from django.db.models.functions import NullIf
        from apps.parts.models import InspectionTemplate

        # Annotate each session with the operation name from the matching
        # InspectionTemplate (part + inspection_type) — same ORM pattern as
        # part_number/part_name. NullIf converts blank names to NULL so the
        # serializer's None-check works correctly.
        template_name_subquery = Subquery(
            InspectionTemplate.objects.filter(
                part=OuterRef('part'),
                inspection_type=OuterRef('inspection_type'),
                is_active=True,
            ).order_by('-version').annotate(
                safe_name=NullIf('name', Value(''))
            ).values('safe_name')[:1]
        )

        qs = InspectionSession.objects.select_related(
            'part', 'machine', 'operator', 'supervisor', 'finalized_by', 'template'
        ).annotate(template_name=template_name_subquery)

        status_filter   = self.request.query_params.get('status')
        machine_code    = self.request.query_params.get('machine')
        part_number     = self.request.query_params.get('part')
        shift           = self.request.query_params.get('shift')
        inspection_type = self.request.query_params.get('inspection_type')
        operator_id     = self.request.query_params.get('operator')
        operator_name   = self.request.query_params.get('operator_name')
        inspector_name  = self.request.query_params.get('inspector_name')
        only_completed  = self.request.query_params.get('only_completed')
        start_date      = self.request.query_params.get('start_date')
        end_date        = self.request.query_params.get('end_date')

        if status_filter:
            qs = qs.filter(status=status_filter)
        if machine_code:
            qs = qs.filter(machine__machine_code__icontains=machine_code)
        if part_number:
            qs = qs.filter(part__part_number__icontains=part_number)
        if shift:
            qs = qs.filter(shift=shift)
        if inspection_type:
            qs = qs.filter(inspection_type=inspection_type)
        if operator_id:
            qs = qs.filter(operator_id=operator_id)
        if operator_name:
            qs = qs.filter(
                Q(operator__username__icontains=operator_name) |
                Q(operator__first_name__icontains=operator_name) |
                Q(operator__last_name__icontains=operator_name)
            )
        if inspector_name:
            qs = qs.filter(
                Q(finalized_by__username__icontains=inspector_name) |
                Q(finalized_by__first_name__icontains=inspector_name) |
                Q(finalized_by__last_name__icontains=inspector_name) |
                Q(supervisor__username__icontains=inspector_name) |
                Q(supervisor__first_name__icontains=inspector_name) |
                Q(supervisor__last_name__icontains=inspector_name)
            )
        if only_completed and only_completed.lower() in ['true', '1', 'yes']:
            qs = qs.filter(
                Q(status__in=['approved', 'finalized_passed', 'completed']) |
                Q(is_setup_approved=True) |
                Q(is_first_piece_finalized=True)
            )
        if start_date:
            qs = qs.filter(started_at__date__gte=start_date)
        if end_date:
            qs = qs.filter(started_at__date__lte=end_date)

        # Filter by plant if user belongs to a specific plant
        if request := self.request:
            if hasattr(request.user, 'plant') and request.user.plant:
                qs = qs.filter(machine__plant=request.user.plant)

        return qs.order_by('-started_at')


# ─── Operator Rejections List ─────────────────────────────────────────────
class RejectionsListView(generics.ListAPIView):
    """
    GET /api/inspections/rejections/
    Returns active rejected sessions that require corrective trial #2 or #3.
    """
    serializer_class   = InspectionSessionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = InspectionSession.objects.select_related(
            'part', 'machine', 'operator', 'supervisor'
        ).filter(
            status=InspectionSession.Status.REJECTED,
            trial_number__lt=3,
        ).order_by('-reviewed_at')

        user = self.request.user
        if hasattr(user, 'plant') and user.plant:
            plant_qs = qs.filter(machine__plant=user.plant)
            if plant_qs.exists():
                return plant_qs
        return qs


# ─── Supervisor 3rd Trial Direct Override ─────────────────────────────────
class SupervisorOverrideView(APIView):
    """
    POST /api/inspections/<session_id>/supervisor-override/
    Allows Supervisor to directly enter/correct a parameter reading on 1ST PC #3.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, session_id):
        if not (request.user.is_supervisor or request.user.is_staff):
            return Response({'error': 'Only supervisors can perform direct overrides.'}, status=status.HTTP_403_FORBIDDEN)

        parameter_code = request.data.get('parameter_code')
        override_value = request.data.get('measured_value')
        remark         = request.data.get('remark', '')

        if not parameter_code or override_value is None:
            return Response({'error': 'parameter_code and measured_value are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            val = float(override_value)
            res = _service.supervisor_override_measurement(
                session_id=session_id,
                parameter_code=parameter_code,
                override_value=val,
                supervisor=request.user,
                remark=remark,
            )
            return Response(res, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# ─── Hourly Time-Lock Status ──────────────────────────────────────────────
class HourlyStatusView(APIView):
    """
    GET /api/inspections/<session_id>/hourly-status/
    Returns open/locked/overdue status for hourly slots 1/HR through 8/HR.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, session_id):
        try:
            res = _service.get_hourly_status(session_id)
            return Response(res, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# ─── Check Setup Approval Status ──────────────────────────────────────────
class SetupStatusView(APIView):
    """
    GET /api/inspections/setup-status/?machine=2
    Returns whether 1st Piece Inspection is approved for a machine.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        machine_id = request.query_params.get('machine')
        if not machine_id:
            return Response({'error': 'machine parameter is required.'}, status=status.HTTP_400_BAD_REQUEST)

        from django.db.models import Q
        from django.utils import timezone
        today = timezone.localdate()

        # Check for today's active session first
        session = None
        if str(machine_id).isdigit():
            session = InspectionSession.objects.select_related('part', 'machine').filter(
                Q(machine_id=int(machine_id)) | Q(machine__machine_code=machine_id),
                started_at__date=today
            ).order_by('-started_at').first()
        else:
            session = InspectionSession.objects.select_related('part', 'machine').filter(
                machine__machine_code=machine_id,
                started_at__date=today
            ).order_by('-started_at').first()

        has_today = True

        # Fallback to latest session if today's check finds no session
        if not session:
            has_today = False
            if str(machine_id).isdigit():
                session = InspectionSession.objects.select_related('part', 'machine').filter(
                    Q(machine_id=int(machine_id)) | Q(machine__machine_code=machine_id)
                ).order_by('-started_at').first()
            else:
                session = InspectionSession.objects.select_related('part', 'machine').filter(
                    machine__machine_code=machine_id
                ).order_by('-started_at').first()

        if not session:
            return Response({
                'has_today_report': False,
                'is_setup_approved': False,
                'status': 'no_session_today',
                'message': 'No inspection started for today.'
            })

        completed_slots = []
        try:
            doc = _service.get_session_detail(str(session.session_id))
            if doc and 'measurements' in doc:
                slot_set = set()
                for m in doc['measurements']:
                    slot = m.get('hourly_slot')
                    if slot and isinstance(slot, int):
                        slot_set.add(slot)
                    elif m.get('inspection_type') == 'hourly' and m.get('hourly_slot'):
                        try:
                            slot_set.add(int(m.get('hourly_slot')))
                        except Exception:
                            pass
                completed_slots = sorted(list(slot_set))
        except Exception:
            pass

        next_slot = (max(completed_slots) + 1) if completed_slots else 1

        is_approved = True

        part_id = session.part.id if session.part else None
        part_no = session.part.part_number if session.part else None
        part_name = session.part.part_name if session.part else None
        machine_id = session.machine.id if session.machine else None

        return Response({
            'has_today_report': has_today,
            'is_setup_approved': is_approved,
            'session_id': str(session.session_id),
            'status': session.status,
            'machine_id': machine_id,
            'part_id': part_id,
            'part_number': part_no,
            'part_name': part_name,
            'inspection_type': session.inspection_type,
            'completed_hourly_slots': completed_slots,
            'next_unlocked_slot': next_slot,
            'message': 'Today\'s inspection report active.' if has_today else 'No inspection started for today.'
        })


# ─── Finalize First Piece Inspection ──────────────────────────────────────
class FinalizeFirstPieceView(APIView):
    """
    POST /api/inspections/<session_id>/finalize/
    Inspector finalizes First Piece inspection and generates PDF.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, session_id):
        try:
            session = _service.finalize_first_piece_session(session_id, request.user)
            return Response(InspectionSessionSerializer(session).data)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# ─── First Piece PDF Report View ──────────────────────────────────────────
class FirstPiecePDFView(APIView):
    """
    GET /api/inspections/<session_id>/pdf/
    Returns the official First Piece Inspection Report PDF file.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, session_id):
        import os
        from django.conf import settings
        from django.http import FileResponse

        try:
            session = InspectionSession.objects.get(session_id=session_id)
            doc = _service.get_session_document(session_id) or {}
            
            pdf_path = session.pdf_report_path
            if not pdf_path or not os.path.exists(os.path.join(settings.BASE_DIR, pdf_path)):
                from .pdf_generator import generate_first_piece_pdf
                pdf_path = generate_first_piece_pdf(session, doc)
                session.pdf_report_path = pdf_path
                session.save(update_fields=['pdf_report_path'])

            abs_pdf_path = os.path.join(settings.BASE_DIR, pdf_path)
            if not os.path.exists(abs_pdf_path):
                return Response({'error': 'PDF report file not found.'}, status=status.HTTP_404_NOT_FOUND)

            return FileResponse(open(abs_pdf_path, 'rb'), content_type='application/pdf', filename=f"FirstPiece_Report_{session_id}.pdf")
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# ─── First Piece Authorization Status View ────────────────────────────────
class FirstPieceStatusView(APIView):
    """
    GET /api/inspections/first-piece-status/?machine_id=2&part_number=PN-101
    Checks if 1st Piece Inspection is finalized and passed for a machine/part.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        machine_id = request.query_params.get('machine_id') or request.query_params.get('machine')
        part_number = request.query_params.get('part_number')

        if not machine_id:
            return Response({'error': 'machine_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        qs = InspectionSession.objects.filter(machine_id=machine_id, inspection_type='first_piece')
        if part_number:
            qs = qs.filter(part__part_number=part_number)

        passed_session = qs.filter(status=InspectionSession.Status.FINALIZED_PASSED).order_by('-finalized_at').first()

        if passed_session:
            return Response({
                'is_authorized': True,
                'status': 'finalized_passed',
                'session_id': str(passed_session.session_id),
                'pdf_url': f"/api/inspections/{passed_session.session_id}/pdf/",
                'message': 'First Piece Inspection is finalized and PASSED. Production authorized.'
            })

        latest_session = qs.order_by('-started_at').first()
        return Response({
            'is_authorized': False,
            'status': latest_session.status if latest_session else 'not_started',
            'session_id': str(latest_session.session_id) if latest_session else None,
            'message': 'First Piece Inspection has not been finalized as PASSED yet.'
        })


# ─── Clear Inspection History (Clean Slate) ──────────────────────────────
class ClearHistoryView(APIView):
    """
    DELETE /api/inspections/clear-history/?machine_code=CNC-01
    or DELETE /api/inspections/clear-history/?session_id=...
    Supervisors can clear live monitoring view for a machine.
    Finalized reports (First Piece, Setup Approval, Production) remain permanent in database.
    """
    permission_classes = [IsSupervisorOrAbove]

    def delete(self, request):
        machine_code = request.query_params.get('machine_code')
        session_id   = request.query_params.get('session_id')

        if session_id:
            session = InspectionSession.objects.filter(session_id=session_id).first()
            if session:
                session.delete()
                _service.collection.delete_one({'_id': str(session_id)})
            return Response({'message': f'Session {session_id} deleted.'})

        if machine_code:
            # Clear ALL sessions for this machine unconditionally to provide a completely clean slate
            active_sessions = InspectionSession.objects.filter(
                machine__machine_code=machine_code
            )
            session_ids = [str(s.session_id) for s in active_sessions]
            active_sessions.delete()
            if session_ids:
                _service.collection.delete_many({'_id': {'$in': session_ids}})

            return Response({
                'message': f'All history completely cleared for machine {machine_code}.'
            })

        return Response({'error': 'machine_code or session_id parameter is required.'}, status=status.HTTP_400_BAD_REQUEST)


# ─── Setup Approval ────────────────────────────────────────────────────────
class SetupApprovalView(APIView):
    """
    POST /api/inspections/setup-approval/
        Inspector submits Setup Approval with Process Parameter readings for 1PC#1, 1PC#2, 1PC#3.
        Stored as a separate document in MongoDB (inspection_type='setup_approval').
        Architecturally SEPARATE from the normal First PC Inspection sessions.

    GET  /api/inspections/setup-approval/?template=<id>&machine=<id>
        Returns the most recent Setup Approval data for a template + machine combination.
        Used to pre-populate the SetupApprovalScreen on the mobile app.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        template_id = request.query_params.get('template')
        machine_id  = request.query_params.get('machine')

        if not template_id or not machine_id:
            return Response(
                {'error': 'Both template and machine query parameters are required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        doc = _service.collection.find_one(
            {
                'inspection_type': 'setup_approval',
                'template_id': int(template_id),
                'machine_id': int(machine_id),
            },
            sort=[('submitted_at', -1)],
        )

        if not doc:
            return Response(
                {'detail': 'No setup approval data found for this template and machine.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        doc['_id'] = str(doc['_id'])
        return Response(doc)

    def post(self, request):
        """
        Payload:
        {
            "template_id": 5,
            "machine_id": 3,
            "part_number": "FBT00222",
            "inspector_name": "Samruddhi Bartakke",
            "process_param_entries": [
                {
                    "parameter_code": "PR1",
                    "parameter_name": "RPM",
                    "trial_1": "1200",
                    "trial_2": "1250",
                    "trial_3": "1200"
                }
            ]
        }
        """
        from datetime import datetime, timezone as tz
        import uuid

        template_id    = request.data.get('template_id')
        machine_id     = request.data.get('machine_id')
        part_number    = request.data.get('part_number', '')
        entries        = request.data.get('process_param_entries', [])
        inspector_name = request.data.get('inspector_name', request.user.get_full_name())

        if not template_id or not machine_id:
            return Response(
                {'error': 'template_id and machine_id are required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not isinstance(entries, list):
            return Response(
                {'error': 'process_param_entries must be a list.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        now = datetime.now(tz.utc)
        doc_id = str(uuid.uuid4())

        doc = {
            '_id': doc_id,
            'inspection_type': 'setup_approval',
            'template_id': int(template_id),
            'machine_id': int(machine_id),
            'part_number': part_number,
            'inspector_id': request.user.id,
            'inspector_name': inspector_name,
            'process_param_entries': entries,
            'submitted_at': now,
            'status': 'submitted',
        }

        # Upsert — update today's existing document or insert new
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        existing = _service.collection.find_one({
            'inspection_type': 'setup_approval',
            'template_id': int(template_id),
            'machine_id': int(machine_id),
            'submitted_at': {'$gte': today_start},
        })

        if existing:
            _service.collection.update_one(
                {'_id': existing['_id']},
                {'$set': {
                    'process_param_entries': entries,
                    'inspector_id': request.user.id,
                    'inspector_name': inspector_name,
                    'submitted_at': now,
                    'status': 'submitted',
                }},
            )
            result_id = str(existing['_id'])
            updated = True
        else:
            _service.collection.insert_one(doc)
            result_id = doc_id
            updated = False

        # Sync process_param_entries to active sessions for this machine & template
        from apps.inspections.models import InspectionSession
        active_sessions = InspectionSession.objects.filter(
            machine_id=int(machine_id),
            started_at__date=now.date(),
        )
        session_ids = [str(s.session_id) for s in active_sessions]
        if session_ids:
            _service.collection.update_many(
                {'_id': {'$in': session_ids}},
                {'$set': {'process_param_entries': entries}}
            )

        return Response(
            {
                'success': True,
                'message': 'Setup Approval process parameters submitted successfully.',
                'setup_approval_id': result_id,
                'template_id': template_id,
                'machine_id': machine_id,
                'submitted_at': now.isoformat(),
                'entries_count': len(entries),
                'updated': updated,
            },
            status=status.HTTP_201_CREATED,
        )


# ─── Daily Production Reports ──────────────────────────────────────────────
class DailyProductionReportViewSet(viewsets.ModelViewSet):
    """
    ViewSet for End-of-Day Daily Production Reports.
    GET /api/inspections/daily-production-reports/
    POST /api/inspections/daily-production-reports/
    GET /api/inspections/daily-production-reports/{id}/export_pdf/
    GET /api/inspections/daily-production-reports/export_excel/
    """
    serializer_class = DailyProductionReportSerializer

    def get_permissions(self):
        if self.action in ['export_excel', 'export_pdf']:
            return []
        return [IsAuthenticated()]

    def get_queryset(self):
        qs = DailyProductionReport.objects.select_related('machine', 'part', 'operator').all()

        date_str = self.request.query_params.get('date')
        if date_str:
            qs = qs.filter(date=date_str)

        machine_id = self.request.query_params.get('machine') or self.request.query_params.get('machine_id')
        if machine_id:
            qs = qs.filter(machine_id=machine_id)

        part_id = self.request.query_params.get('part') or self.request.query_params.get('part_id')
        if part_id:
            qs = qs.filter(part_id=part_id)

        shift = self.request.query_params.get('shift')
        if shift:
            qs = qs.filter(shift=shift)

        operator_id = self.request.query_params.get('operator') or self.request.query_params.get('operator_id')
        if operator_id:
            qs = qs.filter(operator_id=operator_id)

        return qs

    def perform_create(self, serializer):
        # Set operator to current user if not explicitly passed
        operator = serializer.validated_data.get('operator') or self.request.user
        serializer.save(operator=operator)

    @action(detail=True, methods=['get'])
    def export_pdf(self, request, pk=None):
        report = self.get_object()
        relative_path = generate_daily_production_pdf(report)
        full_path = os.path.join(settings.BASE_DIR, relative_path)
        if os.path.exists(full_path):
            return FileResponse(open(full_path, 'rb'), content_type='application/pdf', filename=f"DailyProduction_Report_{report.report_id}.pdf")
        return Response({"error": "PDF generation failed"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        qs = self.filter_queryset(self.get_queryset())
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="Daily_Production_Reports.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'Date', 'Machine Code', 'Machine Name', 'Part Number', 'Part Name', 'Operation',
            'Shift', 'Operator', 'Target', 'Completed', 'Correct', 'Incorrect',
            'CR', 'MR', 'RW', 'Achievement %', 'Remarks'
        ])

        for r in qs:
            op_name = r.operator.get_full_name().strip() if r.operator else ''
            if not op_name and r.operator:
                op_name = r.operator.username
            writer.writerow([
                str(r.date), r.machine.machine_code, r.machine.name,
                r.part.part_number, r.part.part_name, r.operation,
                r.shift, op_name, r.production_target, r.jobs_completed,
                r.correct_jobs, r.incorrect_jobs, r.cr_count, r.mr_count,
                r.rw_count, f"{r.achievement_percentage}%", r.remarks
            ])

        return response


def generate_downtime_xlsx(qs, date_str: str, shift_str: str) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Downtime Report"
    ws.views.sheetView[0].showGridLines = True

    cyan_fill = PatternFill(start_color="B0E0E6", end_color="B0E0E6", fill_type="solid")
    light_blue_fill = PatternFill(start_color="93C5FD", end_color="93C5FD", fill_type="solid")
    sky_fill = PatternFill(start_color="7DD3FC", end_color="7DD3FC", fill_type="solid")
    soft_cyan_fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    title_font = Font(name="Arial", size=13, bold=True)
    header_font = Font(name="Arial", size=20, bold=True)
    sub_font = Font(name="Arial", size=10, bold=True)
    bold_cell_font = Font(name="Arial", size=9, bold=True)
    regular_font = Font(name="Arial", size=9)
    note_font = Font(name="Arial", size=9, italic=True)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # 1. Title Banner (Rows 1 to 3)
    ws.merge_cells("A1:C2")
    ws["A1"] = "HANUMAN ENGINEERING\nWORKS"
    ws["A1"].font = title_font
    ws["A1"].alignment = align_center

    ws.merge_cells("A3:C3")
    ws["A3"] = f"Date: {date_str if date_str and date_str != 'All' else ''}"
    ws["A3"].font = sub_font
    ws["A3"].alignment = align_left

    ws.merge_cells("D1:S3")
    ws["D1"] = "DOWN TIME REPORT"
    ws["D1"].font = header_font
    ws["D1"].alignment = align_center

    ws.merge_cells("T1:V1")
    ws["T1"] = "FORMAT NO. :- QF/MF-06"
    ws["T1"].font = sub_font
    ws["T1"].alignment = align_right

    ws.merge_cells("T2:V2")
    ws["T2"] = "REV. No./ Date :- 00 / 30.09.2026"
    ws["T2"].font = sub_font
    ws["T2"].alignment = align_right

    ws.merge_cells("T3:V3")
    ws["T3"] = f"Shift: {shift_str if shift_str and shift_str != 'All' else ''}"
    ws["T3"].font = sub_font
    ws["T3"].alignment = align_left

    for row in range(1, 4):
        for col in range(1, 23):
            cell = ws.cell(row=row, column=col)
            cell.fill = cyan_fill
            cell.border = thin_border

    # 2. Merged Headers (Rows 4 & 5)
    ws.merge_cells("A4:A5")
    ws["A4"] = "Sr. No."

    ws.merge_cells("B4:B5")
    ws["B4"] = "Machine\nNo."

    ws.merge_cells("C4:C5")
    ws["C4"] = "Operator Name"

    ws.merge_cells("D4:D5")
    ws["D4"] = "Target"

    ws.merge_cells("E4:E5")
    ws["E4"] = "Produced"

    ws.merge_cells("F4:F5")
    ws["F4"] = "Accepted\n/ Actual"

    ws.merge_cells("G4:I4")
    ws["G4"] = "Rejection Summary"
    ws["G5"] = "CR"
    ws["H5"] = "MR"
    ws["I5"] = "RW"

    ws.merge_cells("J4:R4")
    ws["J4"] = "DOWN TIME IN MINUTES"
    ws["J5"] = "NO LOAD"
    ws["K5"] = "NO\nOPERATOR"
    ws["L5"] = "U/M"
    ws["M5"] = "SETTING"
    ws["N5"] = "INSP.\nWAIT"
    ws["O5"] = "TOOL\nCHANGE"
    ws["P5"] = "P/O"
    ws["Q5"] = "R/W"
    ws["R5"] = "TOOL\nPROB"

    ws.merge_cells("S4:S5")
    ws["S4"] = "Total Down\nTime (Min.)"

    ws.merge_cells("T4:V5")
    ws["T4"] = "Remarks"

    for row in (4, 5):
        for col in range(1, 23):
            cell = ws.cell(row=row, column=col)
            cell.font = bold_cell_font
            cell.alignment = align_center
            cell.border = thin_border
            if 7 <= col <= 9:
                cell.fill = light_blue_fill
            elif 10 <= col <= 18:
                cell.fill = sky_fill
            elif col == 19:
                cell.fill = soft_cyan_fill
            else:
                cell.fill = cyan_fill

    # 3. Data Rows (Row 6 onwards)
    current_row = 6
    for idx, obj in enumerate(qs, 1):
        prod = obj.production_report
        op_name = prod.operator.get_full_name().strip() if prod.operator else '—'
        if not op_name and prod.operator:
            op_name = prod.operator.username

        ws.cell(row=current_row, column=1, value=idx).alignment = align_center
        ws.cell(row=current_row, column=2, value=prod.machine.machine_code).alignment = align_center
        ws.cell(row=current_row, column=3, value=op_name).alignment = align_left
        ws.cell(row=current_row, column=4, value=prod.production_target).alignment = align_center
        ws.cell(row=current_row, column=5, value=prod.jobs_completed).alignment = align_center
        ws.cell(row=current_row, column=6, value=prod.correct_jobs).alignment = align_center

        ws.cell(row=current_row, column=7, value=prod.cr_count).alignment = align_center
        ws.cell(row=current_row, column=8, value=prod.mr_count).alignment = align_center
        ws.cell(row=current_row, column=9, value=prod.rw_count).alignment = align_center

        ws.cell(row=current_row, column=10, value=obj.no_load).alignment = align_center
        ws.cell(row=current_row, column=11, value=obj.no_operator).alignment = align_center
        ws.cell(row=current_row, column=12, value=obj.um).alignment = align_center
        ws.cell(row=current_row, column=13, value=obj.setting).alignment = align_center
        ws.cell(row=current_row, column=14, value=obj.inspection_wait).alignment = align_center
        ws.cell(row=current_row, column=15, value=obj.tool_change).alignment = align_center
        ws.cell(row=current_row, column=16, value=obj.power_off).alignment = align_center
        ws.cell(row=current_row, column=17, value=obj.rework).alignment = align_center
        ws.cell(row=current_row, column=18, value=obj.tool_problem).alignment = align_center

        ws.cell(row=current_row, column=19, value=f"=SUM(J{current_row}:R{current_row})").alignment = align_center
        ws.cell(row=current_row, column=19).font = bold_cell_font

        ws.merge_cells(start_row=current_row, start_column=20, end_row=current_row, end_column=22)
        ws.cell(row=current_row, column=20, value=obj.remarks or '').alignment = align_left

        for col in range(1, 23):
            c = ws.cell(row=current_row, column=col)
            c.border = thin_border
            if not c.font.bold:
                c.font = regular_font

        current_row += 1

    # Empty grid rows up to row 19 matching template layout
    while current_row < 20:
        ws.cell(row=current_row, column=1, value=current_row - 5).alignment = align_center
        ws.cell(row=current_row, column=19, value=f"=SUM(J{current_row}:R{current_row})").alignment = align_center
        ws.cell(row=current_row, column=19).font = bold_cell_font
        ws.merge_cells(start_row=current_row, start_column=20, end_row=current_row, end_column=22)

        for col in range(1, 23):
            c = ws.cell(row=current_row, column=col)
            c.border = thin_border
            if not c.font.bold:
                c.font = regular_font
        current_row += 1

    # 4. Total Breakdown Row (Row 20)
    total_row = current_row
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=18)
    ws.cell(row=total_row, column=1, value="Total breakdown Time(Min.)").alignment = align_right
    ws.cell(row=total_row, column=1).font = bold_cell_font

    ws.cell(row=total_row, column=19, value=f"=SUM(S6:S{total_row-1})").alignment = align_center
    ws.cell(row=total_row, column=19).font = bold_cell_font
    ws.cell(row=total_row, column=19).fill = soft_cyan_fill

    ws.merge_cells(start_row=total_row, start_column=20, end_row=total_row, end_column=22)
    for col in range(1, 23):
        ws.cell(row=total_row, column=col).border = thin_border

    # 5. Supervisor Sign Row (Row 21)
    sig_row = total_row + 1
    ws.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row + 1, end_column=22)
    ws.cell(row=sig_row, column=1, value="Supervisor Sign").alignment = Alignment(horizontal="left", vertical="top")
    ws.cell(row=sig_row, column=1).font = Font(name="Arial", size=11, bold=True)
    for r in range(sig_row, sig_row + 2):
        for c in range(1, 23):
            ws.cell(row=r, column=c).border = thin_border

    # 6. Footer Note (Row 23)
    note_row = sig_row + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=18)
    ws.cell(row=note_row, column=1, value="Note: Actual Time to be Recorded in time (minutes)").alignment = align_left
    ws.cell(row=note_row, column=1).font = note_font

    ws.merge_cells(start_row=note_row, start_column=19, end_row=note_row, end_column=22)
    ws.cell(row=note_row, column=19, value="Checked By").alignment = align_right
    ws.cell(row=note_row, column=19).font = bold_cell_font

    for col in range(1, 23):
        ws.cell(row=note_row, column=col).border = thin_border

    # Set Column Widths
    col_widths = {
        'A': 7, 'B': 12, 'C': 16, 'D': 9, 'E': 9, 'F': 11,
        'G': 6, 'H': 6, 'I': 6,
        'J': 9, 'K': 12, 'L': 7, 'M': 9, 'N': 8, 'O': 9, 'P': 6, 'Q': 6, 'R': 8,
        'S': 15, 'T': 10, 'U': 10, 'V': 10
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ─── Downtime Reports ──────────────────────────────────────────────────────
class DowntimeReportViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Downtime Reports attached to Daily Production Reports.
    GET  /api/inspections/downtime-reports/?date=...&shift=...
    POST /api/inspections/downtime-reports/
    POST /api/inspections/downtime-reports/bulk_save/
    GET  /api/inspections/downtime-reports/export_excel/?date=...&shift=...
    GET  /api/inspections/downtime-reports/export_pdf/?date=...&shift=...
    """
    serializer_class = DowntimeReportSerializer

    def get_permissions(self):
        if self.action in ['export_excel', 'export_pdf']:
            return []
        return [IsAuthenticated()]

    def get_queryset(self):
        # Fetch only SUBMITTED Daily Production Reports
        prod_qs = DailyProductionReport.objects.filter(status='SUBMITTED').select_related('machine', 'operator')

        date_str = self.request.query_params.get('date')
        if date_str:
            prod_qs = prod_qs.filter(date=date_str)

        shift = self.request.query_params.get('shift')
        if shift:
            prod_qs = prod_qs.filter(shift=shift)

        machine_id = self.request.query_params.get('machine') or self.request.query_params.get('machine_id')
        if machine_id:
            prod_qs = prod_qs.filter(machine_id=machine_id)

        # Ensure a DowntimeReport exists for each submitted production report
        for prod in prod_qs:
            DowntimeReport.objects.get_or_create(production_report=prod)

        return DowntimeReport.objects.filter(
            production_report__in=prod_qs
        ).select_related(
            'production_report',
            'production_report__machine',
            'production_report__operator',
            'created_by'
        ).order_by('production_report__date', 'production_report__machine__machine_code')

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['post'])
    def bulk_save(self, request):
        data_list = request.data if isinstance(request.data, list) else request.data.get('reports', [])
        if not isinstance(data_list, list):
            return Response({"error": "Expected a list of downtime report records."}, status=status.HTTP_400_BAD_REQUEST)

        updated_reports = []
        for item in data_list:
            prod_id = item.get('production_report_id') or item.get('production_report')
            report_obj = None

            if item.get('id'):
                report_obj = DowntimeReport.objects.filter(id=item['id']).first()

            if not report_obj and prod_id:
                prod = DailyProductionReport.objects.filter(id=prod_id).first()
                if prod:
                    report_obj, _ = DowntimeReport.objects.get_or_create(production_report=prod)

            if not report_obj:
                continue

            serializer = DowntimeReportSerializer(report_obj, data=item, partial=True)
            if serializer.is_valid():
                obj = serializer.save(created_by=request.user)
                if item.get('mark_completed'):
                    obj.status = DowntimeReport.Status.COMPLETED
                    obj.save()
                updated_reports.append(serializer.data)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        return Response(updated_reports, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        qs = self.filter_queryset(self.get_queryset())
        date_str = request.query_params.get('date', 'All')
        shift_str = request.query_params.get('shift', 'All')

        excel_buffer = generate_downtime_xlsx(qs, date_str, shift_str)
        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Downtime_Report_{date_str}_{shift_str}.xlsx"'
        return response

    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        qs = self.filter_queryset(self.get_queryset())
        date_str = request.query_params.get('date', 'All')
        shift_str = request.query_params.get('shift', 'All')

        relative_path = generate_downtime_pdf(qs, date_str, shift_str)
        full_path = os.path.join(settings.BASE_DIR, relative_path)
        if os.path.exists(full_path):
            return FileResponse(
                open(full_path, 'rb'),
                content_type='application/pdf',
                filename=f"Downtime_Report_{date_str}_{shift_str}.pdf"
            )
        return Response({"error": "PDF generation failed"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def history(self, request):
        """
        Returns date-wise and shift-wise summary of submitted downtime reports for history tracking.
        """
        reports = DowntimeReport.objects.select_related(
            'production_report',
            'production_report__machine',
            'created_by'
        ).order_by('-production_report__date')

        date_groups = {}
        for r in reports:
            prod = r.production_report
            if not prod:
                continue
            d_str = str(prod.date)
            shift = prod.shift
            key = f"{d_str}_{shift}"

            if key not in date_groups:
                date_groups[key] = {
                    'key': key,
                    'date': d_str,
                    'shift': shift,
                    'count': 0,
                    'completed_count': 0,
                    'total_downtime': 0,
                    'status': 'COMPLETED' if r.status == 'COMPLETED' else 'PENDING',
                    'updated_at': r.updated_at,
                    'submitted_by': (r.created_by.get_full_name() or r.created_by.username) if r.created_by else 'Supervisor',
                    'machines': []
                }
            
            date_groups[key]['count'] += 1
            if r.status == 'COMPLETED':
                date_groups[key]['completed_count'] += 1
            date_groups[key]['total_downtime'] += r.total_downtime
            if prod.machine and prod.machine.machine_code not in date_groups[key]['machines']:
                date_groups[key]['machines'].append(prod.machine.machine_code)

        history_list = list(date_groups.values())
        return Response(history_list, status=status.HTTP_200_OK)







